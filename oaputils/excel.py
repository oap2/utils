from openpyxl import load_workbook
from openpyxl import Workbook
from io import BytesIO
import os


class Excel(object):
    def __init__(self):
        pass

    def read_from_excel(self, fp: str = None, fo: object = None) -> list:
        """Read from excel file or file object

        Keyword Arguments:
            fp {str} -- excel file path (default: {None})
            fo {object} -- excel file object (default: {None})

        Returns:
            list -- [{'sheet':sheet_name, 'title':[sheet_title], 'data':[sheet_data]}]
        """
        file_data = []
        if fp is not None:
            wb = load_workbook(fp)
        if fo is not None:
            wb = load_workbook(filename=BytesIO(fo.read()))
        for sheet in wb.sheetnames:
            sheet_data = {'sheet': sheet}
            data_sheet = wb[sheet]
            title = [c.value for c in list(data_sheet.rows)[0]]
            sheet_data['title'] = title
            data = []
            for r in list(data_sheet.rows)[1:]:
                data.append([c.value for c in r])
            sheet_data['data'] = data
            file_data.append(sheet_data)
        return file_data

    def write_to_excel(self, file_name: str, headers: list, data: dict, path: str) -> str:
        """Write content to excel file

        Arguments:
            file_name {str} -- excel file name
            headers {list} -- execl sheet headers (['name', 'age'])
            data {dict} -- execl sheet data ({'stu1':[['will','18'],['linda','19']], 'stu2':[['will','18'],['linda','18']])
            path {list} -- execl file path to save

        Returns:
            str -- path/file_name
        """
        wb = Workbook()
        for sheet_name, sheet_data in data.items():
            new_sheet = wb.create_sheet(sheet_name)
            new_sheet.append(headers)
            for row_data in sheet_data:
                new_sheet.append(row_data)
        wb.remove_sheet(wb['Sheet'])
        out_file_name = os.path.join(path, file_name)
        wb.save(out_file_name)
        return out_file_name
